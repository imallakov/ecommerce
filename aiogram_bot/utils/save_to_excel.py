from openpyxl import load_workbook

import os

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from datetime import datetime

from aiogram_bot.bot import bot_send_error_message
from database.models import CartItem
from database.requests import clear_all_cartitems_of_user


async def save_order_to_excel_file(user_id, username, phone_number, items: list[CartItem], address) -> bool:
    try:
        excel_file = "files/orders.xlsx"
        if os.path.exists(excel_file):
            workbook = load_workbook("files/orders.xlsx")
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet['A1'] = 'DATE'
            worksheet['B1'] = 'TIME'
            worksheet['C1'] = 'USER ID'
            worksheet['D1'] = 'USERNAME'
            worksheet['E1'] = 'PHONE NUMBER'
            worksheet['F1'] = 'ORDERED ITEMS'
            worksheet['G1'] = 'ADDRESS'
            header_font = Font(bold=True)
            for row in worksheet["A1:G1"]:
                for cell in row:
                    cell.font = header_font
                    
        now = datetime.now()
        order_str = ', '.join([f"{item.product.name}(x{item.quantity})" for item in items])
        new_row = [
            now.strftime("%d/%m/%y"),
            now.strftime('%H:%M:%S'),
            user_id,
            username,
            phone_number,
            order_str,
            address
        ]
        worksheet.append(new_row)

        directory = 'files'
        if not os.path.exists(directory):
            os.makedirs(directory)

        workbook.save("files/orders.xlsx")
        return await clear_all_cartitems_of_user(user_id=user_id)
    except Exception as error:
        await bot_send_error_message(
            f'save_order_to_excel_file:\nuser_id={user_id}\nusername={username}\nError: {error.__str__()}')
        return False